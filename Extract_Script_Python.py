import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sqlalchemy import create_engine 
import time
import os
from prefect import flow, task

# change the divide range to how many rows of data you want to execute for each batch
# change Connected String and Excel Name
# change table name as per desired choice
# Thats it !!! Way to go
# Update the python modules before using this script, Might end up in some random errors

@flow
def starter():
    connected_string = 'postgresql+psycopg2://postgres:5388@localhost:5432/postgres' #change
    connected_db = create_engine(connected_string)
    Excel_Name = 'Source File - Excel.xlsx' #change to ur desired excel name, must be in same folder
    os.system('cls')
    print("Started Executing !!!")
    time.sleep(2)
    try:
        connected_wb = load_workbook(Excel_Name)
        connected_ws = connected_wb.worksheets[0]
    except:
        print("\nExcel Sheet Connection Error, Retry !!!\n")

    total_rows = connected_ws.max_row
    print("Total Rows : ",total_rows)
    del connected_wb
    del connected_ws

    divide_range = 100 #change this batch value, data will be loaded and stored by batch-by-batch
    table_name ='Supermarket01' #change to ur desired table name
    divide_value = 0
    extract_value = divide_value
    batch_number = 0

    if (total_rows > divide_range):
        while total_rows > divide_range:        
            batch_number = batch_number + 1
            print("\nExecuting Batch : ", batch_number)
            file_fd = pd.read_excel(Excel_Name, skiprows=extract_value +1 , nrows= divide_range, header=None)
            null_rows = [index for index, row in file_fd.iterrows() if row.isnull().any()]
            if null_rows:
                print("Fix null values and try again !!!")
                print("Null Row Indexes are : ", null_rows)
                raise ValueError("Excel sheet has null data in Batch : ", batch_number)
            file_fd.to_sql(table_name,connected_db,if_exists='append', index=False)
            total_rows = total_rows - divide_range
            extract_value = extract_value + divide_range
            del file_fd
            time.sleep(2)

        if(total_rows > 0):
            batch_number = batch_number + 1
            print("\nExecuting Batch : ", batch_number)
            file_fd = pd.read_excel(Excel_Name, skiprows=extract_value +1 , nrows= total_rows, header=None)
            null_rows = [index for index, row in file_fd.iterrows() if row.isnull().any()]
            if null_rows:
                print("Fix null values and try again !!!")
                print("Null Row Indexes are : ", null_rows)
                raise ValueError("Excel sheet has null data in Batch : ", batch_number)
            file_fd.to_sql(table_name,connected_db,if_exists='append',  index=False)
            del file_fd
            time.sleep(2)
    else:
        print("\nExecuting Batch : ", batch_number)
        batch_number = batch_number + 1
        file_fd = pd.read_excel(Excel_Name)
        null_rows = [index for index, row in file_fd.iterrows() if row.isnull().any()]
        if null_rows:
            print("Fix null values and try again !!!")
            print("Null Row Indexes are : ", null_rows)
            raise ValueError("Excel sheet has null data in Batch : ", batch_number)
        file_fd.to_sql(table_name,connected_db,if_exists='append', index=False)
        del file_fd
        time.sleep(2)

    print("\nAll Batches (",batch_number,") has been SQL'ed to the",table_name,"table !!!\n")

if __name__ == "__main__":
    starter.serve(name="To_Postgres_SQL_Deployment", tags=['Scheduled'], interval=780)